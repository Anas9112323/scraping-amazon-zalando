#!/usr/bin/env python3
"""
Export du CSV master vers un Excel (.xlsx) bien formaté.
Ouvre automatiquement le fichier après export.
"""

import sys
import subprocess
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from config import RESULTS_CSV, DATA_DIR


def export():
    if not RESULTS_CSV.exists():
        print("Pas de données. Lance d'abord : python3 pipeline/run_batch.py")
        return

    df = pd.read_csv(RESULTS_CSV, encoding="utf-8-sig")

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx_path = DATA_DIR / f"Leads_Marques_FR_{ts}.xlsx"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_leads = df[df["LEAD"] == "OUI"].copy()
        df_all = df.copy()

        df_leads.to_excel(writer, sheet_name="LEADS", index=False)
        df_all.to_excel(writer, sheet_name="Toutes les marques", index=False)

    wb = load_workbook(xlsx_path)

    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    lead_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    non_lead_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    cell_font = Font(name="Calibri", size=10)
    border = Border(
        left=Side(style="thin", color="dddddd"),
        right=Side(style="thin", color="dddddd"),
        top=Side(style="thin", color="dddddd"),
        bottom=Side(style="thin", color="dddddd"),
    )

    for ws in wb.worksheets:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        lead_col = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == "LEAD":
                lead_col = col_idx
                break

        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            if lead_col:
                val = ws.cell(row=row_idx, column=lead_col).value
                fill = lead_fill if val == "OUI" else non_lead_fill
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        col_widths = {
            "Brand": 22, "Mot-clé": 28, "Amazon présent": 14,
            "Amazon détail": 45, "Amazon note": 10, "Amazon avis": 12,
            "Amazon prix": 14, "Zalando présent (neuf)": 18,
            "Zalando type": 14, "Zalando détail": 45,
            "Site web": 35, "Page contact": 40, "Page RGPD": 40,
            "LEAD": 8, "Date scan": 20,
        }
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=1, column=col_idx).value
            width = col_widths.get(header_val, 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        ws.sheet_properties.tabColor = "16a34a" if "LEAD" in ws.title else "6366f1"

    wb.save(xlsx_path)

    print(f"\n{'='*60}")
    print(f"  EXCEL EXPORTÉ")
    print(f"{'='*60}")
    print(f"  Fichier : {xlsx_path}")
    print(f"  Onglets :")
    print(f"    - LEADS            : {len(df[df['LEAD']=='OUI'])} marques")
    print(f"    - Toutes les marques : {len(df)} marques")
    print(f"\n  Prochaine étape :")
    print(f"  1. Ouvre ce fichier")
    print(f"  2. Upload sur Google Drive")
    print(f"  3. 'Ouvrir avec' → Google Sheets")
    print(f"  4. Partage le lien avec ton collègue")
    print(f"{'='*60}\n")

    try:
        subprocess.run(["open", str(xlsx_path)], check=False)
    except Exception:
        pass

    return xlsx_path


if __name__ == "__main__":
    export()
